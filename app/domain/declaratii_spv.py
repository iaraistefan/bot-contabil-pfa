"""
Generator FISA DE COMPLETARE pentru Decontul special de TVA (D301).

NU genereaza PDF-ul inteligent ANAF (acela necesita programul de asistenta
ANAF). In schimb produce o fisa clara cu valorile EXACTE pe care
contribuabilul le transcrie in formularul oficial - elimina calculul manual.

Caz acoperit: PFA neplatitor de TVA, inregistrat special art. 317, care
primeste servicii intracomunitare (ex. comision Bolt din Estonia). Pentru
aceste servicii se aplica taxarea inversa: se datoreaza TVA 21% pe valoarea
comisionului, fara drept de deducere (fiind neplatitor) - deci TVA datorata
= TVA de plata.
"""

LUNI = {
    1: "Ianuarie", 2: "Februarie", 3: "Martie", 4: "Aprilie",
    5: "Mai", 6: "Iunie", 7: "Iulie", 8: "August",
    9: "Septembrie", 10: "Octombrie", 11: "Noiembrie", 12: "Decembrie",
}

# Cota TVA aplicabila incepand cu 1 august 2025 (Legea 141/2025)
COTA_TVA_CURENTA = 0.21


def construieste_fisa_d301(an: int, luna: int, baza_servicii_intra: float,
                           cota: float = COTA_TVA_CURENTA) -> dict:
    """
    Construieste datele fisei D301 pentru o luna.

    baza_servicii_intra = valoarea serviciilor intracomunitare primite
    (comisionul retinut de platforma in luna respectiva).
    """
    baza = round(baza_servicii_intra, 2)
    tva_datorata = round(baza * cota, 2)
    return {
        "an": an,
        "luna": luna,
        "luna_nume": LUNI.get(luna, str(luna)),
        "cota_pct": round(cota * 100),
        "baza": baza,
        "tva_datorata": tva_datorata,
        "tva_deductibila": 0.0,        # neplatitor - fara drept de deducere
        "tva_de_plata": tva_datorata,  # = TVA datorata
        "termen": f"25 {LUNI.get(luna % 12 + 1, '')}",
    }


def cota_tva_pentru(an: int, luna: int) -> float:
    """Cota TVA aplicabila: 21% de la 1 august 2025, 19% inainte."""
    if an > 2025 or (an == 2025 and luna >= 8):
        return 0.21
    return 0.19


def construieste_fisa_d301_din_tva(an: int, luna: int, tva_de_plata: float) -> dict:
    """
    Construieste fisa pornind de la TVA-ul deja calculat de bot
    (vat_out_total), derivand baza cu cota corecta a perioadei.
    """
    cota = cota_tva_pentru(an, luna)
    baza = round(tva_de_plata / cota, 2) if cota else 0.0
    d = construieste_fisa_d301(an, luna, baza, cota=cota)
    # pastram exact TVA-ul venit din bot (evitam diferente de rotunjire)
    d["tva_datorata"] = round(tva_de_plata, 2)
    d["tva_de_plata"] = round(tva_de_plata, 2)
    return d


def format_fisa_d301(d: dict, profil: dict = None) -> str:
    """
    Fisa de completare structurata EXACT ca formularul D301, sectiune cu
    sectiune, ca sa poata fi transcrisa 1:1 in PDF-ul oficial ANAF.

    Valorile din coloanele de baza/TVA se rotunjesc la leu (asa lucreaza
    formularul). Suma de control si Nr. evidenta platii se genereaza automat
    de PDF la validare - nu se completeaza manual.
    """
    p = profil or {}
    baza_r = round(d["baza"])
    tva_r = round(d["tva_datorata"])

    L = []
    L.append("📄 *DECONT SPECIAL DE TVA - 301*")
    L.append(f"*Pentru luna {d['luna']:02d}  anul {d['an']}*")
    L.append("(Declaratie rectificativa: NU)")
    L.append("===================================")
    L.append("")
    L.append("*1) DATELE DE IDENTIFICARE*")
    L.append(f"• Cod identificare fiscala: {p.get('firma_cui') or '—'}")
    L.append(f"• Denumire / Nume: {p.get('firma_nume') or '—'}")
    adr = p.get("adresa") or p.get("domiciliu") or "—"
    L.append(f"• Adresa: {adr}")
    L.append(f"• Telefon: {p.get('telefon') or '—'}")
    L.append(f"• Banca: {p.get('banca') or '—'}")
    L.append(f"• Cont (IBAN): {p.get('cont') or '—'}")
    L.append("")
    L.append("*2) TIP PERSOANA*")
    L.append("• [X] Persoane inregistrate conform art. 317")
    L.append("")
    L.append("*3) REZUMAT DECLARATIE*")
    L.append("• Suma de control: (auto, la validare)")
    L.append("• Nr. evidenta platii: (auto, la validare)")
    L.append("                    Baza imp.   TVA datorat")
    L.append("  Sectiunea 1          0           0")
    L.append("  Sectiunea 2          0           0")
    L.append("  Sectiunea 3          0           0")
    L.append(f"  Sectiunea 4        {baza_r:>4}        {tva_r:>4}")
    L.append(f"  Sectiunea 4.1      {baza_r:>4}        {tva_r:>4}")
    L.append("")
    L.append("*4) SECTIUNEA 4.1 - detaliu factura*")
    L.append("_Achizitii intracom. de servicii (taxare inversa)_")
    L.append(f"  1. Document Nr/Data: [nr. factura comision] / "
             f"{_ultima_zi(d['an'], d['luna'])}")
    L.append(f"  2. Valoare in valuta: {d['baza']:.2f}")
    L.append("  3. Tip valuta: RON")
    L.append("  4. Curs de schimb: 1")
    L.append(f"  5. Baza de impozitare: {baza_r}")
    L.append(f"  6. TVA datorat: {tva_r}")
    L.append("")
    L.append("  ⚠️ Apasa apoi butonul din formular:")
    L.append("  *Adauga facturi din sectiunea 4.1 in sectiunea 4*")
    L.append("")
    L.append("*5) DECLARATIE PE PROPRIA RASPUNDERE*")
    nume = (p.get("firma_nume") or "").upper()
    L.append(f"• Nume / Prenume: {nume or '—'}")
    L.append("• Functia: TITULAR PFA")
    L.append("")
    L.append(f"💰 *DE PLATA catre ANAF: {tva_r} lei*")
    L.append(f"🗓️ Termen depunere si plata: {d['termen']} "
             f"{d['an'] if d['luna'] < 12 else d['an']+1}")
    L.append("")
    L.append("_In PDF: completeaza, apasa VALIDARE, semneaza, depune in SPV. "
             "Se coreleaza cu D390 (cod S)._")
    return "\n".join(L)


def _ultima_zi(an: int, luna: int) -> str:
    """Ultima zi a lunii, format dd/mm/yyyy (data uzuala a facturii de comision)."""
    import calendar
    zi = calendar.monthrange(an, luna)[1]
    return f"{zi:02d}/{luna:02d}/{an}"


# ============================================================
#          GENERARE XLSX - aspect ca formularul D301
# ============================================================

def genereaza_xlsx_d301(d: dict, profil: dict = None) -> bytes:
    """
    Genereaza un fisier Excel (.xlsx) care reproduce aspectul declaratiei
    D301, completat cu datele lunii. Returneaza continutul ca bytes.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    p = profil or {}
    baza_r = round(d["baza"])
    tva_r = round(d["tva_datorata"])

    wb = Workbook()
    ws = wb.active
    ws.title = f"D301 {d['luna']:02d}-{d['an']}"

    # Stiluri reutilizabile
    f_titlu = Font(name="Arial", size=16, bold=True)
    f_sub = Font(name="Arial", size=11, bold=True)
    f_norm = Font(name="Arial", size=10)
    f_bold = Font(name="Arial", size=10, bold=True)
    f_mic = Font(name="Arial", size=8, italic=True)
    sectiune_fill = PatternFill("solid", start_color="D9E1F2")
    head_fill = PatternFill("solid", start_color="BDD7EE")
    total_fill = PatternFill("solid", start_color="E2EFDA")
    galben = PatternFill("solid", start_color="FFF2CC")
    centru = Alignment(horizontal="center", vertical="center")
    stanga = Alignment(horizontal="left", vertical="center", wrap_text=True)
    dreapta = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Latimi coloane (A..G)
    latimi = {"A": 6, "B": 26, "C": 16, "D": 11, "E": 11, "F": 14, "G": 14}
    for col, w in latimi.items():
        ws.column_dimensions[col].width = w

    r = 1
    # --- Antet ---
    ws.merge_cells(f"A{r}:E{r}")
    ws[f"A{r}"] = "DECONT SPECIAL DE TAXA PE VALOAREA ADAUGATA"
    ws[f"A{r}"].font = f_sub
    ws[f"F{r}"] = "301"
    ws[f"F{r}"].font = f_titlu
    ws[f"F{r}"].alignment = centru
    ws.merge_cells(f"F{r}:G{r}")
    r += 1
    ws.merge_cells(f"A{r}:G{r}")
    ws[f"A{r}"] = f"pentru luna {d['luna']:02d}    anul {d['an']}    (Rectificativa: NU)"
    ws[f"A{r}"].font = f_norm
    r += 2

    # --- Date identificare ---
    ws.merge_cells(f"A{r}:G{r}")
    ws[f"A{r}"] = "DATELE DE IDENTIFICARE A PERSOANEI IMPOZABILE"
    ws[f"A{r}"].font = f_sub
    ws[f"A{r}"].fill = sectiune_fill
    r += 1

    def rand_info(eticheta, valoare):
        nonlocal r
        ws[f"A{r}"] = eticheta
        ws[f"A{r}"].font = f_bold
        ws.merge_cells(f"B{r}:G{r}")
        ws[f"B{r}"] = valoare or "—"
        ws[f"B{r}"].font = f_norm
        ws[f"B{r}"].alignment = stanga
        r += 1

    rand_info("Cod fiscal", p.get("firma_cui"))
    rand_info("Denumire / Nume", p.get("firma_nume"))
    rand_info("Adresa", p.get("adresa") or p.get("domiciliu"))
    rand_info("Telefon", p.get("telefon"))
    rand_info("Banca", p.get("banca"))
    rand_info("Cont (IBAN)", p.get("cont"))
    r += 1

    ws.merge_cells(f"A{r}:G{r}")
    ws[f"A{r}"] = "[X] Persoane inregistrate conform art. 317 din Codul fiscal"
    ws[f"A{r}"].font = f_bold
    r += 2

    # --- Rezumat declaratie ---
    ws.merge_cells(f"A{r}:G{r}")
    ws[f"A{r}"] = "REZUMAT DECLARATIE"
    ws[f"A{r}"].font = f_sub
    ws[f"A{r}"].fill = sectiune_fill
    r += 1
    ws[f"A{r}"] = "Suma de control"
    ws.merge_cells(f"B{r}:C{r}")
    ws[f"B{r}"] = "(auto, la validare)"
    ws[f"B{r}"].font = f_mic
    r += 1
    ws[f"A{r}"] = "Nr. evidenta platii"
    ws.merge_cells(f"B{r}:C{r}")
    ws[f"B{r}"] = "(auto, la validare)"
    ws[f"B{r}"].font = f_mic
    r += 1

    # cap de tabel sectiuni
    ws[f"E{r}"] = "Sectiune"
    ws[f"F{r}"] = "Baza impozitare"
    ws[f"G{r}"] = "TVA datorat"
    for col in ("E", "F", "G"):
        ws[f"{col}{r}"].font = f_bold
        ws[f"{col}{r}"].fill = head_fill
        ws[f"{col}{r}"].alignment = centru
        ws[f"{col}{r}"].border = border
    r += 1
    sectiuni = [("Sectiunea 1", 0, 0), ("Sectiunea 2", 0, 0),
                ("Sectiunea 3", 0, 0), ("Sectiunea 4", baza_r, tva_r),
                ("Sectiunea 4.1", baza_r, tva_r)]
    for nume, b, t in sectiuni:
        ws[f"E{r}"] = nume
        ws[f"F{r}"] = b
        ws[f"G{r}"] = t
        for col in ("E", "F", "G"):
            ws[f"{col}{r}"].font = f_norm
            ws[f"{col}{r}"].border = border
            ws[f"{col}{r}"].alignment = dreapta if col != "E" else stanga
        r += 1
    r += 1

    # --- Sectiunea 4.1 detaliu ---
    ws.merge_cells(f"A{r}:G{r}")
    ws[f"A{r}"] = "Sectiunea 4.1 - Achizitii intracomunitare de servicii (taxare inversa)"
    ws[f"A{r}"].font = f_sub
    ws[f"A{r}"].fill = sectiune_fill
    r += 1
    capete = ["Nr", "Document Nr/Data", "Valoare valuta", "Tip valuta",
              "Curs", "Baza impozit.", "TVA datorat"]
    for i, cap in enumerate(capete):
        c = ws.cell(row=r, column=i + 1, value=cap)
        c.font = f_bold
        c.fill = head_fill
        c.alignment = centru
        c.border = border
    r += 1
    rand_factura = [1, f"[nr. factura] / {_ultima_zi(d['an'], d['luna'])}",
                    round(d["baza"], 2), "RON", 1, baza_r, tva_r]
    for i, val in enumerate(rand_factura):
        c = ws.cell(row=r, column=i + 1, value=val)
        c.font = f_norm
        c.border = border
        c.alignment = centru if i in (0, 3, 4) else (dreapta if i >= 2 else stanga)
    rand_det = r
    r += 1
    # TOTAL
    ws.merge_cells(f"A{r}:E{r}")
    ws[f"A{r}"] = "TOTAL"
    ws[f"A{r}"].font = f_bold
    ws[f"A{r}"].alignment = dreapta
    ws[f"F{r}"] = f"=F{rand_det}"
    ws[f"G{r}"] = f"=G{rand_det}"
    for col in ("A", "F", "G"):
        ws[f"{col}{r}"].fill = total_fill
        ws[f"{col}{r}"].border = border
        ws[f"{col}{r}"].font = f_bold
    ws[f"F{r}"].alignment = dreapta
    ws[f"G{r}"].alignment = dreapta
    r += 2

    # --- De plata + termen ---
    ws.merge_cells(f"A{r}:G{r}")
    ws[f"A{r}"] = f"DE PLATA catre ANAF: {tva_r} lei    |    Termen: {d['termen']} {d['an'] if d['luna'] < 12 else d['an']+1}"
    ws[f"A{r}"].font = f_sub
    ws[f"A{r}"].fill = galben
    r += 2

    # --- Declaratie raspundere ---
    nume = (p.get("firma_nume") or "—").upper()
    ws.merge_cells(f"A{r}:G{r}")
    ws[f"A{r}"] = f"Declar pe propria raspundere: {nume}  -  Functia: TITULAR PFA"
    ws[f"A{r}"].font = f_norm
    r += 1
    ws.merge_cells(f"A{r}:G{r}")
    ws[f"A{r}"] = ("Atentie: valorile se transcriu in PDF-ul inteligent D301 (ANAF), "
                  "apoi VALIDARE + semnare + depunere SPV. Se coreleaza cu D390 (cod S).")
    ws[f"A{r}"].font = f_mic
    ws[f"A{r}"].alignment = stanga

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def nume_fisier_d301(an: int, luna: int) -> str:
    return f"D301_{an}_{luna:02d}.xlsx"


# ============================================================
#          GENERARE PDF - aspect ca formularul ANAF D301
# ============================================================

def genereaza_pdf_d301(d: dict, profil: dict = None) -> bytes:
    """
    Genereaza un PDF care reproduce aspectul formularului D301 ANAF,
    completat cu datele lunii. Returneaza continutul ca bytes.
    """
    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm

    p = profil or {}
    baza_r = round(d["baza"])
    tva_r = round(d["tva_datorata"])

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    W, H = A4
    M = 15 * mm  # margine

    NEGRU = (0, 0, 0)
    ALBASTRU = (0.06, 0.15, 0.42)

    def box(x, y, w, h, lw=0.8):
        c.setLineWidth(lw)
        c.rect(x, y, w, h)

    def txt(x, y, s, font="Helvetica", size=9, color=NEGRU):
        c.setFont(font, size)
        c.setFillColorRGB(*color)
        c.drawString(x, y, str(s))

    def txt_c(xc, y, s, font="Helvetica", size=9, color=NEGRU):
        c.setFont(font, size)
        c.setFillColorRGB(*color)
        c.drawCentredString(xc, y, str(s))

    def txt_r(xr, y, s, font="Helvetica", size=9, color=NEGRU):
        c.setFont(font, size)
        c.setFillColorRGB(*color)
        c.drawRightString(xr, y, str(s))

    def casete(x, y, text, n, cell=5.5 * mm, h=6 * mm):
        # caractere in casute individuale (ca pe formular)
        s = str(text)
        for i in range(n):
            cx = x + i * cell
            box(cx, y, cell, h, lw=0.5)
            if i < len(s):
                c.setFont("Helvetica", 9)
                c.setFillColorRGB(*NEGRU)
                c.drawCentredString(cx + cell / 2, y + 1.6 * mm, s[i])

    y = H - M

    # ===== ANTET =====
    box(M, y - 18 * mm, 55 * mm, 18 * mm)
    txt_c(M + 27.5 * mm, y - 8 * mm, "ANAF", "Helvetica-Bold", 20, ALBASTRU)
    txt_c(M + 27.5 * mm, y - 13 * mm, "Agentia Nationala de", "Helvetica", 6, ALBASTRU)
    txt_c(M + 27.5 * mm, y - 15.5 * mm, "Administrare Fiscala", "Helvetica", 6, ALBASTRU)

    txt_c(M + 110 * mm, y - 7 * mm, "DECONT SPECIAL", "Helvetica-Bold", 13)
    txt_c(M + 110 * mm, y - 13 * mm, "DE TAXA PE VALOAREA ADAUGATA", "Helvetica-Bold", 11)
    txt_r(W - M, y - 14 * mm, "301", "Helvetica-Bold", 30)

    y -= 24 * mm
    txt(M, y, "pentru luna", "Helvetica", 9)
    casete(M + 22 * mm, y - 1.5 * mm, f"{d['luna']:02d}", 2)
    txt(M + 38 * mm, y, "anul", "Helvetica", 9)
    casete(M + 48 * mm, y - 1.5 * mm, str(d["an"]), 4)
    box(M + 95 * mm, y - 2 * mm, 4 * mm, 4 * mm)
    txt(M + 101 * mm, y, "Declaratie rectificativa", "Helvetica", 8)

    y -= 12 * mm
    # ===== DATE IDENTIFICARE =====
    box(M, y - 52 * mm, W - 2 * M, 52 * mm)
    txt(M + 2 * mm, y - 5 * mm, "DATELE DE IDENTIFICARE A PERSOANEI IMPOZABILE",
        "Helvetica-Bold", 9)
    yy = y - 11 * mm
    txt(M + 2 * mm, yy, "Cod de identificare fiscala", "Helvetica", 8)
    txt(M + 48 * mm, yy, "RO", "Helvetica-Bold", 9)
    casete(M + 55 * mm, yy - 1.5 * mm, (p.get("firma_cui") or "").replace("RO", "").strip(), 10)
    yy -= 9 * mm
    txt(M + 2 * mm, yy, "Denumire / Nume, Prenume", "Helvetica", 8)
    txt(M + 50 * mm, yy, p.get("firma_nume") or "—", "Helvetica-Bold", 9)
    yy -= 8 * mm
    txt(M + 2 * mm, yy, "Adresa", "Helvetica", 8)
    txt(M + 50 * mm, yy, (p.get("adresa") or "—")[:60], "Helvetica", 8)
    yy -= 8 * mm
    txt(M + 2 * mm, yy, "Telefon", "Helvetica", 8)
    txt(M + 22 * mm, yy, p.get("telefon") or "—", "Helvetica", 8)
    yy -= 8 * mm
    txt(M + 2 * mm, yy, "Banca", "Helvetica", 8)
    txt(M + 22 * mm, yy, p.get("banca") or "—", "Helvetica", 8)
    txt(M + 90 * mm, yy, "Cont", "Helvetica", 8)
    txt(M + 102 * mm, yy, p.get("cont") or "—", "Helvetica", 8)

    y -= 58 * mm
    # tip persoana
    box(M, y - 6 * mm, 4 * mm, 4 * mm)
    txt(M + 6 * mm, y - 5 * mm, "Persoane care nu sunt inregistrate in scopuri de TVA",
        "Helvetica", 8)
    box(M, y - 13 * mm, 4 * mm, 4 * mm)
    txt(M + 1.0 * mm, y - 12.2 * mm, "X", "Helvetica-Bold", 8)
    txt(M + 6 * mm, y - 12 * mm,
        "Persoane inregistrate conform art. 317 din Legea nr. 227/2015",
        "Helvetica", 8)

    y -= 22 * mm
    # ===== REZUMAT DECLARATIE =====
    txt(M, y, "REZUMAT DECLARATIE", "Helvetica-Bold", 10)
    # tabel control stanga
    box(M, y - 16 * mm, 58 * mm, 13 * mm)
    c.line(M, y - 9.5 * mm, M + 58 * mm, y - 9.5 * mm)
    c.line(M + 40 * mm, y - 16 * mm, M + 40 * mm, y - 3 * mm)
    txt(M + 2 * mm, y - 7.5 * mm, "Suma de control", "Helvetica", 8)
    txt(M + 42 * mm, y - 7.5 * mm, "(auto)", "Helvetica-Oblique", 7)
    txt(M + 2 * mm, y - 14 * mm, "Nr. evidenta a platii", "Helvetica", 8)
    txt(M + 42 * mm, y - 14 * mm, "(auto)", "Helvetica-Oblique", 7)

    # tabel sectiuni dreapta
    sx = M + 95 * mm
    txt_c(sx + 18 * mm, y - 1 * mm, "Baza de", "Helvetica-Bold", 7)
    txt_c(sx + 18 * mm, y - 3.5 * mm, "impozitare (lei)", "Helvetica-Bold", 7)
    txt_c(sx + 40 * mm, y - 1 * mm, "TVA", "Helvetica-Bold", 7)
    txt_c(sx + 40 * mm, y - 3.5 * mm, "datorat (lei)", "Helvetica-Bold", 7)
    sectiuni = [("Sectiune 1", 0, 0), ("Sectiune 2", 0, 0), ("Sectiune 3", 0, 0),
                ("Sectiune 4", baza_r, tva_r), ("Sectiune 4.1", baza_r, tva_r)]
    sy = y - 7 * mm
    for nume, b, t in sectiuni:
        txt(sx - 22 * mm, sy + 1 * mm, nume, "Helvetica", 8)
        box(sx + 6 * mm, sy, 24 * mm, 5 * mm, lw=0.5)
        txt_r(sx + 29 * mm, sy + 1.3 * mm, b, "Helvetica", 8)
        box(sx + 32 * mm, sy, 16 * mm, 5 * mm, lw=0.5)
        txt_r(sx + 47 * mm, sy + 1.3 * mm, t, "Helvetica", 8)
        sy -= 6 * mm

    y = sy - 6 * mm
    # ===== SECTIUNEA 4.1 detaliu =====
    txt(M, y, "Sectiunea 4.1 - Achizitii intracomunitare de servicii (taxare inversa)",
        "Helvetica-Bold", 8)
    y -= 4 * mm
    # cap tabel
    cols = [("Nr", 10), ("Document Numar/Data", 48), ("Valoare in valuta", 28),
            ("Tip valuta", 16), ("Curs", 14), ("Baza impozit.", 22), ("TVA datorat", 22)]
    tw = sum(w for _, w in cols) * mm
    x = M
    rh = 7 * mm
    c.setFillColorRGB(0.85, 0.88, 0.95)
    c.rect(M, y - rh, tw, rh, fill=1, stroke=0)
    c.setFillColorRGB(*NEGRU)
    cx = x
    for nume, w in cols:
        box(cx, y - rh, w * mm, rh, lw=0.5)
        txt_c(cx + w * mm / 2, y - rh + 2.5 * mm, nume, "Helvetica-Bold", 6.5)
        cx += w * mm
    y -= rh
    # rand date
    vals = ["1", f"[nr] / {_ultima_zi(d['an'], d['luna'])}", f"{d['baza']:.2f}",
            "RON", "1", str(baza_r), str(tva_r)]
    cx = x
    for (nume, w), v in zip(cols, vals):
        box(cx, y - rh, w * mm, rh, lw=0.5)
        txt_c(cx + w * mm / 2, y - rh + 2.5 * mm, v, "Helvetica", 7)
        cx += w * mm
    y -= rh
    # total
    box(M, y - rh, tw, rh, lw=0.5)
    txt(M + 2 * mm, y - rh + 2.5 * mm, "TOTAL", "Helvetica-Bold", 7)
    txt_c(x + (10 + 48 + 28 + 16 + 14) * mm + 11 * mm, y - rh + 2.5 * mm, baza_r, "Helvetica-Bold", 7)
    txt_c(x + (10 + 48 + 28 + 16 + 14 + 22) * mm + 11 * mm, y - rh + 2.5 * mm, tva_r, "Helvetica-Bold", 7)

    y -= rh + 10 * mm
    # de plata
    c.setFillColorRGB(1, 0.95, 0.75)
    c.rect(M, y - 8 * mm, W - 2 * M, 8 * mm, fill=1, stroke=1)
    c.setFillColorRGB(*NEGRU)
    txt(M + 3 * mm, y - 5.5 * mm,
        f"DE PLATA catre ANAF: {tva_r} lei    Termen: {d['termen']} "
        f"{d['an'] if d['luna'] < 12 else d['an']+1}", "Helvetica-Bold", 10)

    y -= 16 * mm
    nume = (p.get("firma_nume") or "—").upper()
    txt(M, y, f"Declar pe propria raspundere: {nume}   Functia: TITULAR PFA",
        "Helvetica", 8)
    y -= 6 * mm
    txt(M, y, "Document generat ca model de completare. Se transcrie in PDF-ul "
        "inteligent D301 (ANAF), apoi VALIDARE + semnare + depunere SPV.",
        "Helvetica-Oblique", 7)

    c.showPage()
    c.save()
    return buf.getvalue()


def nume_fisier_d301_pdf(an: int, luna: int) -> str:
    return f"D301_{an}_{luna:02d}.pdf"
