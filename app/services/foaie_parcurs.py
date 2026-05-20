"""
Pas 14 — Foaie de Parcurs: jurnal km auto pentru deductibilitatea
cheltuielilor cu combustibilul.

COMPONENTE:
  • Parser text rapid: `parcurs 18.05 240km Bolt Bistrița`
  • Wizard Telegram cu butoane (înregistrare ghidată)
  • Generare Excel "Foaie de parcurs" formatată
  • Smart link: corelație km ↔ cheltuieli combustibil (plauzibilitate consum)

JUSTIFICARE FISCALĂ:
  Cheltuielile cu combustibilul pentru un vehicul folosit în activitate
  necesită documentare (foaie de parcurs) pentru a justifica gradul de
  deductibilitate. Acest modul produce documentul justificativ.

INTEGRARE:
  • bot_contabil.py — buton meniu + namespace callback "parcurs" + comandă text
  • app.models.TripLog (migration 005)

CHANGELOG:
  • v1 (Pas 14): versiune inițială completă
"""

import io as _io
import logging
import re
from datetime import date, datetime
from typing import Dict, List, Optional, Tuple

from telegram import (
    Update, InlineKeyboardButton, InlineKeyboardMarkup,
)
from telegram.ext import ContextTypes

logger = logging.getLogger(__name__)


# ============================================================
#                    CONSTANTE
# ============================================================

BTN_LABEL = "🚗 Foaie de parcurs"

LUNI_LONG = {
    1: "Ianuarie", 2: "Februarie", 3: "Martie", 4: "Aprilie",
    5: "Mai", 6: "Iunie", 7: "Iulie", 8: "August",
    9: "Septembrie", 10: "Octombrie", 11: "Noiembrie", 12: "Decembrie",
}
LUNI_SHORT = {
    1: "Ian", 2: "Feb", 3: "Mar", 4: "Apr", 5: "Mai", 6: "Iun",
    7: "Iul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec",
}

# Praguri plauzibilitate consum combustibil (L/100km)
CONSUM_MIN_NORMAL = 4.0
CONSUM_MAX_NORMAL = 13.0
# Preț mediu motorină RON/L — folosit pentru estimare litri din valoarea bonului.
# Aproximativ; folosit doar pentru un calcul informativ, nu fiscal.
PRET_MEDIU_MOTORINA = 7.5

# Scopuri predefinite pentru wizard
SCOPURI_PREDEFINITE = [
    "Curse Bolt",
    "Curse Uber",
    "Aprovizionare",
    "Deplasare interes afacere",
]


# ============================================================
#                    PARSER TEXT RAPID
# ============================================================

def is_trip_command(text: str) -> bool:
    """Verifică dacă textul începe cu cuvântul-cheie 'parcurs'."""
    if not text:
        return False
    return text.strip().lower().startswith("parcurs")


def parse_trip_text(text: str) -> Dict:
    """
    Parsează un text de tip:
      parcurs 18.05 240km Bolt Bistrița
      parcurs 18.05.2026 240 km curse Bistrița
      parcurs azi 240km
      parcurs 240 km

    Returns dict:
      {ok: bool, trip_date: date, km: float, purpose: str, error: str}
    """
    result = {
        "ok": False, "trip_date": None, "km": None,
        "purpose": None, "error": None,
    }

    if not text:
        result["error"] = "Text gol"
        return result

    # Eliminăm cuvântul-cheie "parcurs"
    body = text.strip()
    body = re.sub(r"^parcurs\s*", "", body, flags=re.IGNORECASE).strip()

    if not body:
        result["error"] = "Lipsesc datele după 'parcurs'"
        return result

    today = date.today()
    trip_date = today

    # 1. Detectăm data — formate: DD.MM, DD.MM.YYYY, "azi", "ieri"
    tokens = body.split()
    consumed_idx = set()

    first = tokens[0].lower() if tokens else ""
    if first == "azi":
        trip_date = today
        consumed_idx.add(0)
    elif first in ("ieri", "ieri."):
        from datetime import timedelta
        trip_date = today - timedelta(days=1)
        consumed_idx.add(0)
    else:
        # Încercăm DD.MM.YYYY sau DD.MM
        m = re.match(r"^(\d{1,2})\.(\d{1,2})(?:\.(\d{2,4}))?$", first)
        if m:
            day = int(m.group(1))
            month = int(m.group(2))
            year = int(m.group(3)) if m.group(3) else today.year
            if year < 100:
                year += 2000
            try:
                trip_date = date(year, month, day)
                consumed_idx.add(0)
            except ValueError:
                result["error"] = f"Dată invalidă: {first}"
                return result

    # 2. Detectăm km — un număr urmat opțional de "km"
    km_value = None
    for i, tok in enumerate(tokens):
        if i in consumed_idx:
            continue
        # "240km" sau "240"
        m = re.match(r"^(\d+(?:[.,]\d+)?)\s*(km)?$", tok.lower())
        if m:
            km_value = float(m.group(1).replace(",", "."))
            consumed_idx.add(i)
            # Dacă următorul token e "km" separat
            if i + 1 < len(tokens) and tokens[i + 1].lower() == "km":
                consumed_idx.add(i + 1)
            break

    if km_value is None:
        result["error"] = "Nu am găsit numărul de km"
        return result

    if km_value <= 0 or km_value > 2000:
        result["error"] = f"Km în afara intervalului plauzibil: {km_value}"
        return result

    # 3. Restul = scop/traseu
    purpose_tokens = [
        tok for i, tok in enumerate(tokens) if i not in consumed_idx
    ]
    purpose = " ".join(purpose_tokens).strip()
    if not purpose:
        purpose = "Curse Bolt"  # default pentru ridesharing

    result.update({
        "ok": True,
        "trip_date": trip_date,
        "km": km_value,
        "purpose": purpose[:255],
    })
    return result


# ============================================================
#                    DB OPERATIONS
# ============================================================

def register_trip(
    session, user_id: int, trip_date: date,
    km: float, purpose: str,
    odometer_start: Optional[int] = None,
    odometer_end: Optional[int] = None,
) -> Optional[int]:
    """Înregistrează o intrare în foaia de parcurs. Returns trip_id sau None."""
    from app.models import TripLog
    try:
        trip = TripLog(
            user_id=user_id,
            trip_date=trip_date,
            km=km,
            purpose=purpose,
            odometer_start=odometer_start,
            odometer_end=odometer_end,
            period_year=trip_date.year,
            period_month=trip_date.month,
        )
        session.add(trip)
        session.commit()
        return trip.id
    except Exception as e:
        session.rollback()
        logger.error(f"register_trip error: {e}")
        return None


def get_trips_for_period(
    session, user_id: int, year: int, month: int,
) -> List:
    """Returnează intrările foii de parcurs pentru o lună."""
    from app.models import TripLog
    try:
        return (
            session.query(TripLog)
            .filter(
                TripLog.user_id == user_id,
                TripLog.period_year == year,
                TripLog.period_month == month,
            )
            .order_by(TripLog.trip_date)
            .all()
        )
    except Exception as e:
        logger.error(f"get_trips_for_period error: {e}")
        return []


def get_available_trip_periods(session, user_id: int) -> List[Tuple[int, int]]:
    """Returnează (year, month) pentru care există intrări."""
    from app.models import TripLog
    try:
        rows = (
            session.query(TripLog.period_year, TripLog.period_month)
            .filter(TripLog.user_id == user_id)
            .distinct()
            .all()
        )
        return sorted(
            set((r[0], r[1]) for r in rows), reverse=True
        )
    except Exception as e:
        logger.error(f"get_available_trip_periods error: {e}")
        return []


def delete_trip(session, user_id: int, trip_id: int) -> bool:
    """Șterge o intrare din foaia de parcurs."""
    from app.models import TripLog
    try:
        trip = (
            session.query(TripLog)
            .filter(TripLog.id == trip_id, TripLog.user_id == user_id)
            .first()
        )
        if not trip:
            return False
        session.delete(trip)
        session.commit()
        return True
    except Exception as e:
        session.rollback()
        logger.error(f"delete_trip error: {e}")
        return False


# ============================================================
#               SMART LINK — CORELAȚIE COMBUSTIBIL
# ============================================================

def _get_fuel_expense_for_period(
    session, user_id: int, year: int, month: int,
) -> float:
    """
    Suma cheltuielilor cu combustibilul pentru o lună (RON).
    Caută documente de tip cheltuială cu cuvinte-cheie combustibil.
    """
    try:
        from app.models import Document
        target_month_str = f"{month:02d}.{year}"
        fuel_keywords = [
            "motorina", "motorină", "benzina", "benzină",
            "combustibil", "carburant", "diesel", "gpl",
            "lukoil", "omv", "petrom", "rompetrol", "mol", "socar",
        ]

        docs = (
            session.query(Document)
            .filter(
                Document.user_id == user_id,
                Document.status == "posted",
            )
            .all()
        )

        total = 0.0
        for d in docs:
            if not d.data_doc or target_month_str not in d.data_doc:
                continue
            text = f"{d.platforma or ''} {d.detalii or ''}".lower()
            if any(kw in text for kw in fuel_keywords):
                total += float(d.brut or 0)

        return round(total, 2)
    except Exception as e:
        logger.error(f"_get_fuel_expense_for_period error: {e}")
        return 0.0


def analyze_fuel_consistency(
    session, user_id: int, year: int, month: int,
) -> Dict:
    """
    Analizează plauzibilitatea: km parcurși vs cheltuieli combustibil.

    Returns dict:
      {
        total_km, fuel_ron, estimated_liters,
        consum_per_100km, verdict, verdict_emoji, message
      }
    """
    trips = get_trips_for_period(session, user_id, year, month)
    total_km = sum(t.km for t in trips)
    fuel_ron = _get_fuel_expense_for_period(session, user_id, year, month)

    result = {
        "total_km": total_km,
        "fuel_ron": fuel_ron,
        "estimated_liters": 0.0,
        "consum_per_100km": 0.0,
        "cost_per_km": 0.0,
        "verdict": "no_data",
        "verdict_emoji": "ℹ️",
        "message": "",
    }

    if total_km <= 0:
        result["message"] = "Nu există km înregistrați pentru această lună."
        return result

    if fuel_ron <= 0:
        result["message"] = (
            "Nu există cheltuieli cu combustibilul înregistrate. "
            "Înregistrează bonurile pentru analiză completă."
        )
        result["cost_per_km"] = 0.0
        return result

    estimated_liters = fuel_ron / PRET_MEDIU_MOTORINA
    consum = (estimated_liters / total_km) * 100.0
    cost_per_km = fuel_ron / total_km

    result["estimated_liters"] = round(estimated_liters, 1)
    result["consum_per_100km"] = round(consum, 1)
    result["cost_per_km"] = round(cost_per_km, 2)

    if CONSUM_MIN_NORMAL <= consum <= CONSUM_MAX_NORMAL:
        result["verdict"] = "ok"
        result["verdict_emoji"] = "✅"
        result["message"] = (
            f"Consum estimat {consum:.1f} L/100km — plauzibil. "
            f"Foaia de parcurs susține cheltuielile cu combustibilul."
        )
    elif consum < CONSUM_MIN_NORMAL:
        result["verdict"] = "low"
        result["verdict_emoji"] = "⚠️"
        result["message"] = (
            f"Consum estimat {consum:.1f} L/100km — neobișnuit de mic. "
            f"Posibil: km supraevaluați sau bonuri combustibil lipsă."
        )
    else:
        result["verdict"] = "high"
        result["verdict_emoji"] = "⚠️"
        result["message"] = (
            f"Consum estimat {consum:.1f} L/100km — peste normal. "
            f"Posibil: km subevaluați sau combustibil pentru alt vehicul."
        )

    return result


# ============================================================
#               GENERARE EXCEL — FOAIE DE PARCURS
# ============================================================

def generate_foaie_parcurs_xlsx(
    trips: List, year: int, month: int,
    pfa_name: str = "PFA", pfa_cui: str = "",
    fuel_analysis: Optional[Dict] = None,
) -> bytes:
    """
    Generează Excel "Foaie de parcurs" formatat, gata de tipărit.
    Returns: bytes XLSX.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = f"Foaie parcurs {LUNI_SHORT.get(month, '')}"

    # Stiluri
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    cell_font = Font(size=10)
    total_font = Font(bold=True, size=11)
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")

    # --- Titlu ---
    ws.merge_cells("A1:F1")
    ws["A1"] = "FOAIE DE PARCURS"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    ws.merge_cells("A2:F2")
    ws["A2"] = f"{LUNI_LONG.get(month, '')} {year}"
    ws["A2"].font = Font(bold=True, size=11)
    ws["A2"].alignment = center

    # --- Date firmă ---
    ws["A4"] = "Titular:"
    ws["A4"].font = Font(bold=True, size=10)
    ws["B4"] = pfa_name
    ws["A5"] = "CUI/CIF:"
    ws["A5"].font = Font(bold=True, size=10)
    ws["B5"] = pfa_cui or "—"

    # --- Header tabel ---
    header_row = 7
    headers = [
        "Nr.", "Data", "Km parcurși",
        "Citire bord\n(start)", "Citire bord\n(sfârșit)",
        "Scop / Traseu",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # --- Rânduri date ---
    row = header_row + 1
    total_km = 0.0
    for idx, trip in enumerate(trips, start=1):
        ws.cell(row=row, column=1, value=idx).alignment = center
        ws.cell(
            row=row, column=2,
            value=trip.trip_date.strftime("%d.%m.%Y"),
        ).alignment = center
        ws.cell(row=row, column=3, value=round(trip.km, 1)).alignment = center
        ws.cell(
            row=row, column=4,
            value=trip.odometer_start if trip.odometer_start else "—",
        ).alignment = center
        ws.cell(
            row=row, column=5,
            value=trip.odometer_end if trip.odometer_end else "—",
        ).alignment = center
        ws.cell(
            row=row, column=6, value=trip.purpose or "—",
        ).alignment = left

        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border
            if ws.cell(row=row, column=col).font.size is None:
                ws.cell(row=row, column=col).font = cell_font

        total_km += trip.km
        row += 1

    # --- Rând TOTAL ---
    ws.cell(row=row, column=1, value="TOTAL").font = total_font
    ws.merge_cells(
        start_row=row, start_column=1, end_row=row, end_column=2
    )
    ws.cell(row=row, column=1).alignment = center
    total_cell = ws.cell(row=row, column=3, value=round(total_km, 1))
    total_cell.font = total_font
    total_cell.alignment = center
    total_cell.fill = PatternFill("solid", fgColor="D9E2F3")
    for col in range(1, 7):
        ws.cell(row=row, column=col).border = border

    # --- Analiză combustibil (opțional) ---
    if fuel_analysis and fuel_analysis.get("verdict") not in ("no_data",):
        info_row = row + 2
        ws.merge_cells(
            start_row=info_row, start_column=1,
            end_row=info_row, end_column=6,
        )
        ws.cell(
            row=info_row, column=1,
            value="ANALIZĂ COMBUSTIBIL (informativ)",
        ).font = Font(bold=True, size=10)

        fa = fuel_analysis
        details = [
            f"Total km luna: {fa['total_km']:.0f} km",
            f"Cheltuieli combustibil: {fa['fuel_ron']:.2f} RON",
            f"Litri estimați: ~{fa['estimated_liters']:.1f} L",
            f"Consum estimat: {fa['consum_per_100km']:.1f} L/100km",
            f"Cost mediu: {fa['cost_per_km']:.2f} RON/km",
        ]
        for i, d in enumerate(details, start=1):
            ws.cell(row=info_row + i, column=1, value=d).font = cell_font
            ws.merge_cells(
                start_row=info_row + i, start_column=1,
                end_row=info_row + i, end_column=6,
            )

    # --- Semnătură ---
    sig_row = row + (10 if fuel_analysis else 3)
    ws.cell(
        row=sig_row, column=1,
        value="Întocmit (semnătură): ____________________",
    ).font = cell_font
    ws.cell(
        row=sig_row, column=5,
        value=f"Data: {date.today().strftime('%d.%m.%Y')}",
    ).font = cell_font

    # --- Lățimi coloane ---
    widths = [6, 14, 14, 14, 14, 40]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Setări print
    ws.print_area = f"A1:F{sig_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def filename_foaie_parcurs(year: int, month: int) -> str:
    return f"Foaie_parcurs_{year}_{month:02d}.xlsx"


# ============================================================
#               UI TELEGRAM — KEYBOARD BUILDERS
# ============================================================

def _build_main_menu() -> InlineKeyboardMarkup:
    """Meniul principal al foii de parcurs."""
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(
            "➕ Adaugă zi de parcurs",
            callback_data="parcurs|add"
        )],
        [InlineKeyboardButton(
            "📊 Vezi luna curentă",
            callback_data="parcurs|view_current"
        )],
        [InlineKeyboardButton(
            "📥 Descarcă foaie Excel",
            callback_data="parcurs|export"
        )],
        [InlineKeyboardButton(
            "⛽ Analiză combustibil",
            callback_data="parcurs|fuel"
        )],
        [InlineKeyboardButton("❌ Închide", callback_data="nav|close")],
    ])


def _build_purpose_picker() -> InlineKeyboardMarkup:
    """Picker pentru scop predefinit."""
    rows = []
    for i in range(0, len(SCOPURI_PREDEFINITE), 2):
        row = [
            InlineKeyboardButton(
                s, callback_data=f"parcurs|purpose|{idx}"
            )
            for idx, s in enumerate(
                SCOPURI_PREDEFINITE[i:i + 2], start=i
            )
        ]
        rows.append(row)
    rows.append([
        InlineKeyboardButton("⬅️ Înapoi", callback_data="parcurs|menu"),
    ])
    return InlineKeyboardMarkup(rows)


def _build_period_picker(
    periods: List[Tuple[int, int]], action: str,
) -> InlineKeyboardMarkup:
    """Picker pentru perioadă (year, month)."""
    rows = []
    for i in range(0, len(periods), 2):
        row = []
        for year, month in periods[i:i + 2]:
            row.append(InlineKeyboardButton(
                f"{LUNI_SHORT.get(month)} {year}",
                callback_data=f"parcurs|{action}|{year}|{month}"
            ))
        rows.append(row)
    rows.append([
        InlineKeyboardButton("⬅️ Înapoi", callback_data="parcurs|menu"),
    ])
    return InlineKeyboardMarkup(rows)


# ============================================================
#               UI TELEGRAM — HANDLERS
# ============================================================

async def handle_menu_button(
    update: Update, context: ContextTypes.DEFAULT_TYPE,
) -> None:
    """Apelat când user apasă butonul '🚗 Foaie de parcurs' din meniu."""
    msg = (
        "🚗 *FOAIE DE PARCURS*\n"
        "━━━━━━━━━━━━━━━━━━━━\n\n"
        "Jurnalul km parcurși — *justifică deductibilitatea "
        "cheltuielilor cu combustibilul.*\n\n"
        "✍️ *Înregistrare rapidă (text):*\n"
        "`parcurs 18.05 240km Bolt Bistrița`\n"
        "`parcurs azi 180 km aprovizionare`\n\n"
        "_Sau folosește butoanele de mai jos:_"
    )
    await update.message.reply_text(
        msg, parse_mode="Markdown",
        reply_markup=_build_main_menu(),
    )


async def handle_text_command(
    update: Update, context: ContextTypes.DEFAULT_TYPE,
) -> bool:
    """
    Procesează comanda text `parcurs ...`.
    Returns True dacă a procesat, False altfel.
    """
    from db import get_session
    from app.repositories import users as users_repo

    text = update.message.text or ""
    if not is_trip_command(text):
        return False

    parsed = parse_trip_text(text)
    if not parsed["ok"]:
        await update.message.reply_text(
            f"⚠️ Nu am putut înregistra parcursul.\n"
            f"_{parsed['error']}_\n\n"
            f"Format corect:\n"
            f"`parcurs 18.05 240km Bolt Bistrița`",
            parse_mode="Markdown",
        )
        return True

    tg_id = update.effective_user.id
    session = get_session()
    try:
        user = users_repo.get_by_telegram_id(session, telegram_id=tg_id)
        if not user:
            await update.message.reply_text("⚠️ Eroare identificare utilizator.")
            return True

        trip_id = register_trip(
            session, user.id,
            trip_date=parsed["trip_date"],
            km=parsed["km"],
            purpose=parsed["purpose"],
        )

        if trip_id:
            # Total luna curentă pentru context
            trips = get_trips_for_period(
                session, user.id,
                parsed["trip_date"].year,
                parsed["trip_date"].month,
            )
            total_km = sum(t.km for t in trips)

            await update.message.reply_text(
                f"✅ *Parcurs înregistrat* #{trip_id}\n\n"
                f"📅 Data: {parsed['trip_date'].strftime('%d.%m.%Y')}\n"
                f"🚗 Km: *{parsed['km']:.0f} km*\n"
                f"📍 Scop: {parsed['purpose']}\n\n"
                f"📊 Total {LUNI_LONG.get(parsed['trip_date'].month)} "
                f"{parsed['trip_date'].year}: *{total_km:.0f} km* "
                f"({len(trips)} zile)",
                parse_mode="Markdown",
            )
        else:
            await update.message.reply_text(
                "❌ Eroare la înregistrarea parcursului."
            )
        return True
    except Exception as e:
        logger.error(f"handle_text_command error: {e}")
        await update.message.reply_text("❌ Eroare sistem.")
        return True
    finally:
        session.close()


async def handle_callback(
    update: Update, context: ContextTypes.DEFAULT_TYPE, parts: List[str],
) -> None:
    """Router pentru callback queries namespace=parcurs."""
    from db import get_session
    from app.repositories import users as users_repo

    query = update.callback_query
    tg_id = update.effective_user.id

    if len(parts) < 2:
        return
    action = parts[1]

    session = get_session()
    try:
        user = users_repo.get_by_telegram_id(session, telegram_id=tg_id)
        if not user:
            await query.edit_message_text("⚠️ Eroare identificare utilizator.")
            return
        user_id = user.id

        # ─── MENIU ───────────────────────────────────────────
        if action == "menu":
            await query.edit_message_text(
                "🚗 *FOAIE DE PARCURS*\n\nAlege o opțiune:",
                parse_mode="Markdown",
                reply_markup=_build_main_menu(),
            )
            return

        # ─── ADAUGĂ (instrucțiuni text) ──────────────────────
        if action == "add":
            await query.edit_message_text(
                "➕ *Adaugă zi de parcurs*\n\n"
                "Trimite un mesaj text în formatul:\n\n"
                "`parcurs 18.05 240km Bolt Bistrița`\n"
                "`parcurs azi 180 km aprovizionare`\n"
                "`parcurs 15.05.2026 220 km`\n\n"
                "_Data poate fi `azi`, `ieri`, sau `ZZ.LL`._\n"
                "_Daca nu specifici scopul, se trece Curse Bolt._",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton(
                        "⬅️ Înapoi", callback_data="parcurs|menu"
                    )
                ]]),
            )
            return

        # ─── VEZI LUNA CURENTĂ ───────────────────────────────
        if action == "view_current":
            today = date.today()
            await _show_period_trips(
                query, session, user_id, today.year, today.month
            )
            return

        # ─── EXPORT — alege perioada ─────────────────────────
        if action == "export":
            periods = get_available_trip_periods(session, user_id)
            if not periods:
                await query.edit_message_text(
                    "📭 Nu ai înregistrări în foaia de parcurs.\n\n"
                    "Adaugă prima zi cu:\n"
                    "`parcurs azi 240km Bolt`",
                    parse_mode="Markdown",
                    reply_markup=InlineKeyboardMarkup([[
                        InlineKeyboardButton(
                            "⬅️ Înapoi", callback_data="parcurs|menu"
                        )
                    ]]),
                )
                return
            await query.edit_message_text(
                "📥 *Descarcă foaia de parcurs*\nAlege luna:",
                parse_mode="Markdown",
                reply_markup=_build_period_picker(periods, "export_do"),
            )
            return

        # ─── EXPORT — generează Excel ────────────────────────
        if action == "export_do":
            year = int(parts[2])
            month = int(parts[3])
            await _generate_and_send_excel(
                query, context, session, user_id, year, month
            )
            return

        # ─── ANALIZĂ COMBUSTIBIL ─────────────────────────────
        if action == "fuel":
            today = date.today()
            analysis = analyze_fuel_consistency(
                session, user_id, today.year, today.month
            )
            msg = _format_fuel_analysis(analysis, today.year, today.month)
            await query.edit_message_text(
                msg, parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton(
                        "⬅️ Înapoi", callback_data="parcurs|menu"
                    )
                ]]),
            )
            return

    except Exception as e:
        logger.error(f"foaie_parcurs callback error: {e}")
        try:
            await query.edit_message_text(f"❌ Eroare: {str(e)[:200]}")
        except Exception:
            pass
    finally:
        session.close()


# ============================================================
#               HELPERS — DISPLAY
# ============================================================

async def _show_period_trips(query, session, user_id, year, month):
    """Afișează intrările pentru o lună."""
    trips = get_trips_for_period(session, user_id, year, month)

    if not trips:
        await query.edit_message_text(
            f"📭 Nu ai înregistrări pentru "
            f"{LUNI_LONG.get(month)} {year}.\n\n"
            f"Adaugă cu: `parcurs azi 240km Bolt`",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton(
                    "⬅️ Înapoi", callback_data="parcurs|menu"
                )
            ]]),
        )
        return

    total_km = sum(t.km for t in trips)
    lines = [
        f"📊 *Foaie parcurs — {LUNI_LONG.get(month)} {year}*",
        "━━━━━━━━━━━━━━━━━━━━",
        "",
    ]
    for t in trips:
        lines.append(
            f"📅 {t.trip_date.strftime('%d.%m')} — "
            f"*{t.km:.0f} km* · _{t.purpose or '—'}_"
        )

    lines.extend([
        "",
        "━━━━━━━━━━━━━━━━━━━━",
        f"🚗 *TOTAL: {total_km:.0f} km* ({len(trips)} zile)",
    ])

    await query.edit_message_text(
        "\n".join(lines),
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton(
                "📥 Descarcă Excel",
                callback_data=f"parcurs|export_do|{year}|{month}"
            )],
            [InlineKeyboardButton(
                "⬅️ Înapoi", callback_data="parcurs|menu"
            )],
        ]),
    )


async def _generate_and_send_excel(
    query, context, session, user_id, year, month,
):
    """Generează și trimite Excel-ul foii de parcurs."""
    from app.repositories import users as users_repo

    await query.edit_message_text(
        f"🔄 Generez foaia de parcurs "
        f"{LUNI_LONG.get(month)} {year}..."
    )

    trips = get_trips_for_period(session, user_id, year, month)
    if not trips:
        await query.edit_message_text(
            f"📭 Nu există înregistrări pentru "
            f"{LUNI_LONG.get(month)} {year}."
        )
        return

    profile = users_repo.get_profile_dict(session, user_id) or {}
    pfa_name = profile.get("firma_nume") or "PFA"
    pfa_cui = profile.get("firma_cui") or ""

    fuel_analysis = analyze_fuel_consistency(
        session, user_id, year, month
    )

    try:
        xlsx_bytes = generate_foaie_parcurs_xlsx(
            trips, year, month,
            pfa_name=pfa_name, pfa_cui=pfa_cui,
            fuel_analysis=fuel_analysis,
        )
        fname = filename_foaie_parcurs(year, month)
        total_km = sum(t.km for t in trips)

        await context.bot.send_document(
            chat_id=query.message.chat_id,
            document=_io.BytesIO(xlsx_bytes),
            filename=fname,
            caption=(
                f"🚗 *Foaie de parcurs — "
                f"{LUNI_LONG.get(month)} {year}*\n\n"
                f"📊 Total: {total_km:.0f} km · {len(trips)} zile\n"
                f"🖨️ Tipărește: Print → Landscape A4\n\n"
                f"_Document justificativ pentru deductibilitatea "
                f"cheltuielilor auto._"
            ),
            parse_mode="Markdown",
        )
        await query.edit_message_text(
            f"✅ Foaie de parcurs generată pentru "
            f"{LUNI_LONG.get(month)} {year}."
        )
    except Exception as e:
        logger.error(f"_generate_and_send_excel error: {e}")
        await query.edit_message_text(
            "❌ Eroare la generarea foii de parcurs."
        )


def _format_fuel_analysis(analysis: Dict, year: int, month: int) -> str:
    """Formatează rezultatul analizei combustibil."""
    lines = [
        f"⛽ *ANALIZĂ COMBUSTIBIL*",
        f"_{LUNI_LONG.get(month)} {year}_",
        "━━━━━━━━━━━━━━━━━━━━",
        "",
    ]

    if analysis["verdict"] == "no_data":
        lines.append(f"ℹ️ {analysis['message']}")
        return "\n".join(lines)

    lines.extend([
        f"🚗 Total km: *{analysis['total_km']:.0f} km*",
        f"💰 Cheltuieli combustibil: *{analysis['fuel_ron']:.2f} RON*",
    ])

    if analysis["fuel_ron"] > 0:
        lines.extend([
            f"⛽ Litri estimați: ~{analysis['estimated_liters']:.1f} L",
            f"📊 Consum estimat: *{analysis['consum_per_100km']:.1f} "
            f"L/100km*",
            f"💵 Cost mediu: {analysis['cost_per_km']:.2f} RON/km",
        ])

    lines.extend([
        "",
        f"{analysis['verdict_emoji']} _{analysis['message']}_",
        "",
        "_Calculul litrilor e estimativ (preț mediu motorină). "
        "Pentru exactitate, păstrează bonurile cu litri afișați._",
    ])

    return "\n".join(lines)


__all__ = [
    "BTN_LABEL",
    "is_trip_command",
    "handle_menu_button",
    "handle_text_command",
    "handle_callback",
]
