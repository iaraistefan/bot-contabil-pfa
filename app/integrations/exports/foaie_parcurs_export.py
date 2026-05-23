"""
Pas A.4 - Generator foaie de parcurs Excel (format ANAF) si calcul
deductibilitate combustibil.

Foaia de parcurs justifica deductibilitatea cheltuielilor auto. Conform
legii trebuie sa cuprinda: identificarea vehiculului, categoria,
scopul/locul deplasarii, km parcursi si norma de consum carburant.

Kilometrajul din foaie e CONTINUU: gap-urile dintre ture (km personali)
apar ca randuri distincte "Utilizare personala", astfel incat foaia sa
fie coerenta la un control.
"""

import io
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models import TRIP_STATUS_CLOSED

logger = logging.getLogger(__name__)

LUNI_LONG = {
    1: "Ianuarie", 2: "Februarie", 3: "Martie", 4: "Aprilie",
    5: "Mai", 6: "Iunie", 7: "Iulie", 8: "August",
    9: "Septembrie", 10: "Octombrie", 11: "Noiembrie", 12: "Decembrie",
}

# Pret de referinta motorina (RON/litru) - estimativ, se poate ajusta.
# Folosit doar pentru estimarea valorii; litrii normati sunt valoarea certa.
PRET_MOTORINA_REFERINTA = 7.5

# --- Stiluri ---
_FONT_TITLU = Font(name="Calibri", size=15, bold=True, color="1F4E78")
_FONT_SUBTITLU = Font(name="Calibri", size=11, bold=True, color="404040")
_FONT_LABEL = Font(name="Calibri", size=10, bold=True, color="404040")
_FONT_VAL = Font(name="Calibri", size=10, color="000000")
_FONT_HEAD = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
_FONT_CELL = Font(name="Calibri", size=10, color="000000")
_FONT_PERSONAL = Font(name="Calibri", size=10, italic=True, color="808080")
_FONT_TOTAL = Font(name="Calibri", size=10, bold=True, color="000000")

_FILL_HEAD = PatternFill("solid", fgColor="1F4E78")
_FILL_TOTAL = PatternFill("solid", fgColor="D9E2F3")
_FILL_PERSONAL = PatternFill("solid", fgColor="F2F2F2")
_FILL_CONSUM = PatternFill("solid", fgColor="E2EFDA")

_thin = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_RIGHT = Alignment(horizontal="right", vertical="center")


# ============================================================
#       CALCUL DEDUCTIBILITATE COMBUSTIBIL
# ============================================================

def calcul_deductibilitate_combustibil(km_business: float,
                                       norma_consum: float,
                                       pret_litru: float = None) -> dict:
    """
    Calculeaza combustibilul aferent activitatii pe baza foii de parcurs.

    Metoda consumului normat (recunoscuta fiscal):
      litri_normati = km_business x norma_consum / 100
      valoare       = litri_normati x pret_litru

    Pentru activitatile exceptate de la plafonul de 50% (ridesharing,
    taxi, curierat), aceasta valoare e integral deductibila.

    Args:
        km_business  : km parcursi in interes business (din foaia parcurs)
        norma_consum : L/100km
        pret_litru   : RON/litru (default: pretul de referinta)

    Returns dict cu detaliul calculului.
    """
    if pret_litru is None:
        pret_litru = PRET_MOTORINA_REFERINTA

    litri_normati = round(km_business * norma_consum / 100.0, 2)
    valoare = round(litri_normati * pret_litru, 2)

    return {
        "km_business": km_business,
        "norma_consum": norma_consum,
        "litri_normati": litri_normati,
        "pret_litru": pret_litru,
        "valoare_deductibila": valoare,
    }


# ============================================================
#       CONSTRUIRE RANDURI (cu gap-uri personale intercalate)
# ============================================================

def _build_rows(trips: list) -> list:
    """
    Construieste randurile foii de parcurs din lista de ture.

    Intercaleaza randuri "Utilizare personala" pentru gap-urile de
    kilometraj, astfel incat odometrul sa fie continuu.

    Fiecare rand: dict cu tip ('business'|'personal') si campuri.
    """
    closed = sorted(
        [t for t in trips if t.status == TRIP_STATUS_CLOSED],
        key=lambda t: (t.trip_date, t.odometer_start or 0),
    )

    rows = []
    prev_end = None
    for t in closed:
        # Gap personal inainte de aceasta tura?
        if (prev_end is not None and t.odometer_start is not None
                and t.odometer_start > prev_end):
            gap = t.odometer_start - prev_end
            rows.append({
                "tip": "personal",
                "data": "",
                "ora_start": "", "ora_stop": "",
                "traseu": "Utilizare personala",
                "odo_start": prev_end,
                "odo_stop": t.odometer_start,
                "km": gap,
            })

        rows.append({
            "tip": "business",
            "data": t.trip_date.strftime("%d.%m.%Y") if t.trip_date else "",
            "ora_start": t.ora_start or "",
            "ora_stop": t.ora_stop or "",
            "traseu": t.purpose or "Curse / activitate",
            "odo_start": t.odometer_start,
            "odo_stop": t.odometer_end,
            "km": t.km or 0,
        })
        if t.odometer_end is not None:
            prev_end = t.odometer_end

    return rows


# ============================================================
#       GENERATOR EXCEL
# ============================================================

def generate_foaie_parcurs_xlsx(trips: list, year: int, month: int,
                                vehicul, pfa_name: str = "PFA",
                                pfa_cui: str = "",
                                pret_litru: float = None) -> bytes:
    """
    Genereaza foaia de parcurs lunara in format Excel.

    Args:
        trips      : lista de TripLog din luna
        year, month: perioada
        vehicul    : obiectul Vehicul (poate fi None)
        pfa_name   : numele PFA/firmei
        pfa_cui    : CUI
        pret_litru : pret motorina pentru estimarea valorii

    Returns: continutul fisierului .xlsx ca bytes.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Parcurs {LUNI_LONG.get(month, '')[:3]} {year}"
    ws.sheet_view.showGridLines = False

    # Latimi coloane (A..H)
    widths = [6, 13, 10, 10, 30, 14, 14, 13]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    LAST_COL = 8  # H
    r = 1

    # --- TITLU ---
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=LAST_COL)
    cell = ws.cell(row=r, column=1, value="FOAIE DE PARCURS")
    cell.font = _FONT_TITLU
    cell.alignment = _CENTER
    ws.row_dimensions[r].height = 26
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=LAST_COL)
    cell = ws.cell(row=r, column=1, value=f"{LUNI_LONG.get(month, '')} {year}")
    cell.font = _FONT_SUBTITLU
    cell.alignment = _CENTER
    r += 2

    # --- BLOC IDENTIFICARE ---
    def _info_row(label, value):
        nonlocal r
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = _FONT_LABEL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        c2 = ws.cell(row=r, column=3, value=value)
        c2.font = _FONT_VAL
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=LAST_COL)
        r += 1

    nr_inmat = vehicul.nr_inmatriculare if vehicul else "—"
    marca = (vehicul.marca_model if vehicul and vehicul.marca_model else "—")
    norma = vehicul.norma_consum if vehicul else 7.5

    _info_row("Titular:", pfa_name)
    if pfa_cui:
        _info_row("CUI / CIF:", pfa_cui)
    _info_row("Autovehicul:", f"{nr_inmat}  ({marca})")
    _info_row("Categoria:", "Autoturism — transport persoane (ridesharing)")
    _info_row("Norma de consum:", f"{norma:g} litri / 100 km")
    r += 1

    # --- HEADER TABEL ---
    headers = [
        "Nr.", "Data", "Plecare", "Sosire", "Scop / Traseu",
        "Km bord\nplecare", "Km bord\nsosire", "Km\nparcursi",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=col, value=h)
        cell.font = _FONT_HEAD
        cell.fill = _FILL_HEAD
        cell.alignment = _CENTER
        cell.border = _BORDER
    ws.row_dimensions[r].height = 30
    r += 1

    # --- RANDURI ---
    rows = _build_rows(trips)
    total_business = 0.0
    total_personal = 0.0
    nr_crt = 0

    for row in rows:
        is_personal = row["tip"] == "personal"
        if is_personal:
            total_personal += row["km"]
            nr_label = ""
        else:
            nr_crt += 1
            total_business += row["km"]
            nr_label = str(nr_crt)

        values = [
            nr_label,
            row["data"],
            row["ora_start"],
            row["ora_stop"],
            row["traseu"],
            row["odo_start"] if row["odo_start"] is not None else "",
            row["odo_stop"] if row["odo_stop"] is not None else "",
            round(row["km"], 1) if row["km"] else 0,
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = _BORDER
            if is_personal:
                cell.font = _FONT_PERSONAL
                cell.fill = _FILL_PERSONAL
            else:
                cell.font = _FONT_CELL
            if col == 5:
                cell.alignment = _LEFT
            elif col in (6, 7, 8):
                cell.alignment = _RIGHT
            else:
                cell.alignment = _CENTER
        r += 1

    # --- RAND TOTAL ---
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    cell = ws.cell(row=r, column=1, value="TOTAL KM PARCURSI IN INTERES BUSINESS")
    cell.font = _FONT_TOTAL
    cell.fill = _FILL_TOTAL
    cell.alignment = _RIGHT
    cell.border = _BORDER
    cell = ws.cell(row=r, column=8, value=round(total_business, 1))
    cell.font = _FONT_TOTAL
    cell.fill = _FILL_TOTAL
    cell.alignment = _RIGHT
    cell.border = _BORDER
    r += 1

    if total_personal > 0:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        cell = ws.cell(row=r, column=1, value="Total km utilizare personala")
        cell.font = _FONT_PERSONAL
        cell.alignment = _RIGHT
        cell.border = _BORDER
        cell = ws.cell(row=r, column=8, value=round(total_personal, 1))
        cell.font = _FONT_PERSONAL
        cell.alignment = _RIGHT
        cell.border = _BORDER
        r += 1

    r += 1

    # --- BLOC CONSUM & DEDUCTIBILITATE ---
    ded = calcul_deductibilitate_combustibil(total_business, norma, pret_litru)

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=LAST_COL)
    cell = ws.cell(row=r, column=1,
                   value="CONSUM CARBURANT AFERENT ACTIVITATII")
    cell.font = _FONT_SUBTITLU
    cell.fill = _FILL_CONSUM
    cell.alignment = _CENTER
    r += 1

    def _consum_row(label, value, bold=False):
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = _FONT_TOTAL if bold else _FONT_VAL
        c1.alignment = _LEFT
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=LAST_COL)
        c2 = ws.cell(row=r, column=6, value=value)
        c2.font = _FONT_TOTAL if bold else _FONT_VAL
        c2.alignment = _RIGHT
        r += 1

    _consum_row("Km parcursi in interes business", f"{round(total_business, 1):g} km")
    _consum_row("Norma de consum", f"{norma:g} L / 100 km")
    _consum_row("Combustibil normat (litri aferenti activitatii)",
                f"{ded['litri_normati']:g} litri", bold=True)
    _consum_row(f"Pret mediu motorina (estimativ {ded['pret_litru']:g} RON/L)",
                f"{ded['pret_litru']:g} RON/L")
    _consum_row("Valoare combustibil aferenta activitatii (estimativ)",
                f"{ded['valoare_deductibila']:.2f} RON", bold=True)
    r += 1

    # --- NOTA ---
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=LAST_COL)
    nota = ws.cell(
        row=r, column=1,
        value=(
            "Nota: Activitatea de transport persoane (ridesharing) este "
            "exceptata de la plafonul de 50% pentru cheltuielile auto. "
            "Combustibilul aferent km business documentati prin aceasta "
            "foaie de parcurs este deductibil. Valoarea in RON este "
            "estimativa - se confrunta cu bonurile fiscale reale. "
            "Confirmati aplicarea cu contabilul."
        ),
    )
    nota.font = Font(name="Calibri", size=8, italic=True, color="808080")
    nota.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    r += 4

    # --- SEMNATURA ---
    ws.cell(row=r, column=1, value="Data intocmirii: ____________").font = _FONT_VAL
    cell = ws.cell(row=r, column=6, value="Semnatura: ____________")
    cell.font = _FONT_VAL
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=LAST_COL)

    # --- Setari print ---
    ws.print_area = f"A1:{get_column_letter(LAST_COL)}{r}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def filename_foaie_parcurs(year: int, month: int, nr_inmatriculare: str = "") -> str:
    """Construieste numele fisierului foii de parcurs."""
    nr = (nr_inmatriculare or "").replace(" ", "").replace("-", "")
    nr_part = f"_{nr}" if nr else ""
    return f"Foaie_parcurs{nr_part}_{year}_{month:02d}.xlsx"
