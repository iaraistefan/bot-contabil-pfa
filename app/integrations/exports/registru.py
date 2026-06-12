"""
Generator Registru de Încasări și Plăți pentru PFA.

Format conform Ordinului MFP 170/2015 (PFA sistem real) — cod 14-1-1/b.

CONȚINUT MINIMAL OBLIGATORIU (per OMFP 170/2015):
1. Data efectuării operațiunii de încasare/plată
2. Denumirea și numărul documentului justificativ
3. Explicații privind natura operațiunii
4. Suma încasată în numerar sau prin bancă
5. Suma plătită în numerar sau prin bancă

STRUCTURĂ COLOANE (7 coloane standard profesionale):
A. Nr. crt.
B. Data
C. Document (Bon fiscal / Factură / Raport / Chitanță)
D. Explicații (natura operațiunii)
E. Încasări (RON)
F. Plăți (RON)
G. Sold cumulat

CHANGELOG:
- v1: Versiune inițială cu 10 coloane (cash/card separate)
- v2: SUMAR FINANCIAR la final
- v3 (cercetare profesională, mai 2026): Format 7 coloane standard OMFP,
  conform modele profesionale (contzilla.ro, lege5.ro, fiscalitatea.ro,
  portalpfa.ro, top-contabilitate.ro). Cash/card NU se mai separă (informația
  e în documentul justificativ și extras de cont). Fix-uri aplicate:
  • Floating point precision: round(sold, 2) la fiecare operațiune
  • Formule SUM consistente pe rândul TOTAL (inclusiv coloana Sold)
  • Footer alignment consistent cu SUMAR (A:D + E:G)
  • Document detection: tip auto-detectat din categorie (Bon fiscal /
    Factură / Raport / Chitanță)

⚠️ MULTI-TENANT (Bug #7 fix):
- Default-urile sunt GENERICE ("PFA", "") — niciun nume hardcodat
- Apelantul TREBUIE să furnizeze pfa_name și pfa_cui din profilul user-ului
"""

import io
import logging
from datetime import date, datetime
from typing import Optional

logger = logging.getLogger(__name__)

LUNI_RO_UPPER = {
    1: "IANUARIE", 2: "FEBRUARIE", 3: "MARTIE", 4: "APRILIE",
    5: "MAI", 6: "IUNIE", 7: "IULIE", 8: "AUGUST",
    9: "SEPTEMBRIE", 10: "OCTOMBRIE", 11: "NOIEMBRIE", 12: "DECEMBRIE"
}


# ============================================================
#                    HELPERS
# ============================================================

def _parse_date(date_str: Optional[str]) -> Optional[date]:
    if not date_str:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(date_str), fmt).date()
        except (ValueError, TypeError):
            continue
    return None


def _calc_row_height(text: str, col_width_chars: int = 48) -> int:
    """Calculează înălțimea optimă a rândului în funcție de lungimea textului."""
    if not text:
        return 22
    chars_per_line = int(col_width_chars * 1.4)
    lines_needed = max(1, (len(text) + chars_per_line - 1) // chars_per_line)
    return max(22, lines_needed * 18)


def _validate_pfa_info(pfa_name: str, pfa_cui: str) -> tuple:
    """Validează și normalizează info PFA."""
    name = (pfa_name or "").strip() or "PFA — Nume nesetat"
    cui = (pfa_cui or "").strip() or ""

    if not pfa_cui:
        logger.warning(
            "generate_registru_xlsx called without pfa_cui — "
            "user profile probably missing"
        )

    return name, cui


def _resolve_document_type(tx) -> str:
    """
    Returnează tipul documentului justificativ conform convențiilor românești.

    Conform OMFP 170/2015, coloana "denumirea documentului justificativ" trebuie
    să indice clar tipul documentului. În modelele profesionale româneşti
    (contzilla, lege5, etc.) se folosesc tipurile standard:
      - "Bon fiscal" — pentru bonuri de la casa de marcat (combustibil, retail)
      - "Factură" — pentru servicii, abonamente, materiale
      - "Raport" — pentru sumarele de venituri (Bolt, Uber, marketplace)
      - "Chitanță" — pentru onorarii notar/avocat/contabil
      - "Poliță" — pentru asigurări
      - "Document oficial" — pentru taxe, autorizații
    """
    cat = tx.category or ""
    cp = (tx.counterparty or "").strip()
    is_real_cp = cp and cp not in ("N/A", "APP", "Platform")

    if tx.tx_type == "INCOME":
        if cat == "ride_revenue":
            return f"Raport {cp}" if is_real_cp else "Raport venituri"
        if cat == "tip_revenue":
            return "Raport bacșișuri"
        if cat == "services_revenue":
            return "Factură emisă"
        return "Document venit"

    # EXPENSE
    if cat == "platform_commission":
        return f"Factură {cp}" if is_real_cp else "Factură comision"
    if cat == "professional_fees":
        return "Chitanță"
    if cat == "car_insurance":
        return "Poliță asigurare"
    if cat == "registration":
        return "Document oficial"
    if cat in ("software_subscriptions", "telecom", "materials", "services"):
        return "Factură"
    # Default pentru cheltuieli fizice (combustibil, service, spălătorie, etc.)
    return "Bon fiscal"


# Etichete generice pe categorii
CATEGORY_LABELS = {
    # Venituri
    "ride_revenue": "Venituri curse",
    "tip_revenue": "Bacșișuri",
    "services_revenue": "Venituri din servicii",
    # Cheltuieli auto
    "fuel": "Combustibil auto",
    "car_service": "Service / Consumabile auto",
    "car_insurance": "Asigurări auto (RCA, CASCO)",
    "car_wash": "Spălătorie auto",
    "car_supplies": "Accesorii auto",
    # Platforme
    "platform_commission": "Comision platformă",
    # Operaționale
    "registration": "Taxe autorizații/înregistrare",
    "professional_fees": "Onorariu notar/contabil/avocat",
    "telecom": "Telefon / Internet",
    "software_subscriptions": "Abonamente software",
    "materials": "Materiale",
    "services": "Servicii (utilități, abonamente)",
    "other_expense": "Alte cheltuieli",
}


def _resolve_category_label(category: Optional[str]) -> str:
    """Returnează eticheta umană pentru o categorie."""
    return CATEGORY_LABELS.get(
        category or "",
        (category or "—").replace("_", " ").title()
    )


def _build_explicatii(tx) -> str:
    """
    Construiește textul pentru coloana Explicații.

    Format:
      - Categorie [+ — Furnizor] [+ (numerar/cont bancar) pentru venituri]
    """
    explicatie = _resolve_category_label(tx.category)

    # Adăugăm furnizorul (counterparty) doar dacă e un nume specific
    cp = (tx.counterparty or "").strip()
    if cp and cp not in ("N/A", "Bolt", "Uber", "APP", "Platform"):
        explicatie += f" — {cp}"

    # Pentru venituri ride/tip, adăugăm metoda de plată (numerar vs cont bancar)
    # Asta face cele 2 rânduri (cash + card) clar distinctive în registru
    if tx.tx_type == "INCOME" and tx.category in ("ride_revenue", "tip_revenue"):
        if tx.payment_method == "CASH":
            explicatie += " (în numerar)"
        elif tx.payment_method == "CARD":
            explicatie += " (prin cont bancar)"

    return explicatie


# ============================================================
#              TOTALURI (sursă unică: Excel + banner)
# ============================================================

def _relevant_txs(transactions, year: int, month=None):
    """Tranzacțiile relevante pt. registru (INCOME/EXPENSE + luna dacă e lunar),
    sortate pe (occurred_on, id). SURSĂ UNICĂ — folosită de Excel ȘI de totaluri."""
    rt = [
        tx for tx in transactions
        if tx.tx_type in ("INCOME", "EXPENSE")
        and (month is None or (tx.occurred_on and tx.occurred_on.month == month))
    ]
    rt.sort(key=lambda tx: (
        tx.occurred_on or date(year, 1, 1),
        tx.id if tx.id else 0,
    ))
    return rt


def registru_totals(transactions, year: int, month=None) -> dict:
    """Totaluri registru — SURSĂ UNICĂ pentru Excel + banner (cifre identice garantat).

    incasari = Σ amount_brut (INCOME), plati = Σ amount_brut (EXPENSE),
    sold = incasari − plati (CU SEMN — poate fi negativ). `last` = ultima înregistrare
    (dată + explicații, trunchiată) sau None dacă nu există tranzacții.
    """
    rt = _relevant_txs(transactions, year, month)
    incasari = round(sum(tx.amount_brut or 0.0 for tx in rt if tx.tx_type == "INCOME"), 2)
    plati = round(sum(tx.amount_brut or 0.0 for tx in rt if tx.tx_type == "EXPENSE"), 2)
    sold = round(incasari - plati, 2)
    last = None
    if rt:
        t = rt[-1]
        zi = t.occurred_on.strftime("%d.%m.%Y") if t.occurred_on else "—"
        expl = _build_explicatii(t)
        if len(expl) > 40:
            expl = expl[:37] + "..."
        last = f"{zi} · {expl}"
    return {"incasari": incasari, "plati": plati, "sold": sold, "last": last}


# ============================================================
#                  MAIN GENERATOR
# ============================================================

def generate_registru_xlsx(
    transactions,
    year: int,
    pfa_name: str = "PFA",
    pfa_cui: str = "",
    month: Optional[int] = None,
) -> bytes:
    """
    Generează Registrul XLSX conform OMFP 170/2015 (cod 14-1-1/b).

    Args:
        transactions: lista de Transaction din DB
        year: anul perioadei
        pfa_name: numele firmei (din profilul user-ului)
        pfa_cui: CUI-ul firmei (din profilul user-ului)
        month: None pentru registru ANUAL, 1..12 pentru registru LUNAR.
    """
    pfa_name, pfa_cui = _validate_pfa_info(pfa_name, pfa_cui)

    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side
        )
    except ImportError:
        logger.error("openpyxl not installed — falling back to CSV")
        return generate_registru_csv(transactions, year, pfa_name, pfa_cui, month)

    wb = openpyxl.Workbook()
    ws = wb.active

    if month:
        ws.title = (
            f"Registru "
            f"{LUNI_RO_UPPER.get(month, str(month))[:3].title()} {year}"
        )
        title_period = f"{LUNI_RO_UPPER.get(month, str(month))} {year}"
    else:
        ws.title = f"Registru {year}"
        title_period = f"ANUL {year}"

    # ── Stiluri ──────────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor="1F3864")
    subheader_fill = PatternFill("solid", fgColor="2E75B6")
    income_fill = PatternFill("solid", fgColor="E2EFDA")
    income_fill_alt = PatternFill("solid", fgColor="F1F8EC")
    expense_fill = PatternFill("solid", fgColor="FCE4D6")
    expense_fill_alt = PatternFill("solid", fgColor="FBE9E7")
    total_fill = PatternFill("solid", fgColor="FFF2CC")
    profit_positive_fill = PatternFill("solid", fgColor="FFF2CC")
    profit_negative_fill = PatternFill("solid", fgColor="FFD7D7")
    month_sep_fill = PatternFill("solid", fgColor="DEEAF1")

    white_bold = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    dark_bold = Font(name="Calibri", bold=True, color="1F3864", size=10)
    normal = Font(name="Calibri", size=10)
    small = Font(name="Calibri", size=9, color="595959")

    center = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    left = Alignment(
        horizontal="left", vertical="center", wrap_text=True
    )
    right = Alignment(horizontal="right", vertical="center")
    left_indent = Alignment(
        horizontal="left", vertical="center", wrap_text=True, indent=2
    )

    thin = Side(style="thin", color="BFBFBF")
    thick = Side(style="medium", color="1F3864")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_border = Border(left=thick, right=thick, top=thick, bottom=thick)

    num_fmt = '#,##0.00 "RON"'

    # ── Lățimi coloane (7 col — format standard OMFP) ────────────────────
    col_widths = {
        "A": 7,    # Nr. crt.
        "B": 13,   # Data
        "C": 20,   # Document
        "D": 48,   # Explicații
        "E": 18,   # Încasări
        "F": 18,   # Plăți
        "G": 18,   # Sold
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    row = 1

    # ── Titlu principal (merged A:G) ─────────────────────────────────────
    ws.merge_cells(f"A{row}:G{row}")
    c = ws[f"A{row}"]
    c.value = f"REGISTRU DE ÎNCASĂRI ȘI PLĂȚI — {title_period}"
    c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
    c.fill = header_fill
    c.alignment = center
    ws.row_dimensions[row].height = 32
    row += 1

    # ── Info PFA ────────────────────────────────────────────────────────
    cui_display = f"CUI: {pfa_cui}" if pfa_cui else "CUI: nesetat"
    ws.merge_cells(f"A{row}:G{row}")
    c = ws[f"A{row}"]
    c.value = (
        f"{pfa_name}  |  {cui_display}  |  "
        f"Sistem real de impunere  |  cod 14-1-1/b"
    )
    c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    c.fill = subheader_fill
    c.alignment = center
    ws.row_dimensions[row].height = 22
    row += 1

    row += 1  # spacing

    # ── Header tabel (7 col) ─────────────────────────────────────────────
    headers = [
        ("A", "Nr.\ncrt."),
        ("B", "Data"),
        ("C", "Document\njustificativ"),
        ("D", "Explicații\n(natura operațiunii)"),
        ("E", "ÎNCASĂRI\n(RON)"),
        ("F", "PLĂȚI\n(RON)"),
        ("G", "SOLD\ncumulat"),
    ]
    for col, label in headers:
        c = ws[f"{col}{row}"]
        c.value = label
        c.font = white_bold
        c.fill = subheader_fill
        c.alignment = center
        c.border = header_border
    ws.row_dimensions[row].height = 42
    row += 1

    # ── Filtrare + sortare tranzacții ────────────────────────────────────
    relevant_txs = _relevant_txs(transactions, year, month)

    sold_curent = 0.0
    nr_crt = 0
    data_start_row = row

    # ── Sold inițial ─────────────────────────────────────────────────────
    sold_label_date = (
        f"01.{month:02d}.{year}" if month else f"01.01.{year}"
    )
    for col in ["A", "B", "C", "D", "E", "F", "G"]:
        c = ws[f"{col}{row}"]
        c.border = thin_border
        c.font = dark_bold
        c.fill = total_fill
    ws.merge_cells(f"C{row}:F{row}")
    ws[f"A{row}"].value = "—"
    ws[f"A{row}"].alignment = center
    ws[f"B{row}"].value = sold_label_date
    ws[f"B{row}"].alignment = center
    ws[f"C{row}"].value = "SOLD INIȚIAL"
    ws[f"C{row}"].alignment = left_indent
    ws[f"G{row}"].value = 0.00
    ws[f"G{row}"].number_format = num_fmt
    ws[f"G{row}"].alignment = right
    ws.row_dimensions[row].height = 24
    row += 1

    current_month = 0

    # ── Tranzacții (rânduri principale) ─────────────────────────────────
    for tx in relevant_txs:
        tx_date = tx.occurred_on or date(year, 1, 1)
        tx_month = tx_date.month

        # Separator lunar (doar pentru registru ANUAL)
        if month is None and tx_month != current_month:
            current_month = tx_month
            ws.merge_cells(f"A{row}:G{row}")
            c = ws[f"A{row}"]
            c.value = (
                f"── {LUNI_RO_UPPER.get(tx_month, str(tx_month))} {year} ──"
            )
            c.font = Font(
                name="Calibri", bold=True, color="2E75B6", size=10
            )
            c.fill = month_sep_fill
            c.alignment = center
            c.border = thin_border
            ws.row_dimensions[row].height = 20
            row += 1

        nr_crt += 1
        is_income = tx.tx_type == "INCOME"

        # FIX BUG #1: rotunjire pentru a evita floating point precision
        amount = round(tx.amount_brut or 0.0, 2)
        if is_income:
            incasare = amount
            plata = 0.0
            sold_curent = round(sold_curent + incasare, 2)
        else:
            incasare = 0.0
            plata = amount
            sold_curent = round(sold_curent - plata, 2)

        # Construim document type + explicații
        document_type = _resolve_document_type(tx)
        explicatie = _build_explicatii(tx)

        row_height = _calc_row_height(explicatie, col_width_chars=48)

        # Colorare fill — alternantă subtilă pentru lizibilitate
        if is_income:
            fill = income_fill_alt if nr_crt % 2 == 0 else income_fill
        else:
            fill = expense_fill_alt if nr_crt % 2 == 0 else expense_fill

        # Populare 7 coloane
        data = {
            "A": (nr_crt, center),
            "B": (tx_date.strftime("%d.%m.%Y"), center),
            "C": (document_type, center),
            "D": (explicatie, left),
            "E": (incasare if incasare else None, right),
            "F": (plata if plata else None, right),
            "G": (sold_curent, right),
        }

        for col, (val, align) in data.items():
            c = ws[f"{col}{row}"]
            c.value = val
            c.alignment = align
            c.font = normal
            c.fill = fill
            c.border = thin_border
            if col in ("E", "F", "G") and val is not None:
                c.number_format = num_fmt
            if col == "G":
                c.font = Font(
                    name="Calibri", bold=True, size=10,
                    color="1F6B2A" if sold_curent >= 0 else "C00000"
                )

        ws.row_dimensions[row].height = row_height
        row += 1

    # ── Total general (formule SUM consistente) ──────────────────────────
    row += 1
    period_label = f"TOTAL {title_period}"

    # Label merged A:D
    ws.merge_cells(f"A{row}:D{row}")
    for col in ["A", "B", "C", "D"]:
        c = ws[f"{col}{row}"]
        c.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        c.fill = header_fill
        c.border = header_border
    ws[f"A{row}"].value = period_label
    ws[f"A{row}"].alignment = Alignment(
        horizontal="left", vertical="center", indent=2
    )

    # E: SUM Încasări
    cE = ws[f"E{row}"]
    cE.value = f"=SUM(E{data_start_row + 1}:E{row - 2})"
    cE.number_format = num_fmt
    cE.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    cE.fill = header_fill
    cE.alignment = right
    cE.border = header_border

    # F: SUM Plăți
    cF = ws[f"F{row}"]
    cF.value = f"=SUM(F{data_start_row + 1}:F{row - 2})"
    cF.number_format = num_fmt
    cF.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    cF.fill = header_fill
    cF.alignment = right
    cF.border = header_border

    # G: FIX BUG #2 — formulă =E-F (consistent cu restul SUM-urilor)
    cG = ws[f"G{row}"]
    cG.value = f"=E{row}-F{row}"
    cG.number_format = num_fmt
    cG.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    cG.fill = header_fill
    cG.alignment = right
    cG.border = header_border

    ws.row_dimensions[row].height = 28
    row += 2

    # ════════════════════════════════════════════════════════════════════
    # ── 📊 SUMAR FINANCIAR (cererea băncii + utilitate contabil) ───────
    # ════════════════════════════════════════════════════════════════════
    # Sursă unică: aceleași totaluri pe care le arată și banner-ul (Registru).
    _tot = registru_totals(transactions, year, month)
    total_incasari_all = _tot["incasari"]
    total_plati_all = _tot["plati"]
    profit_net = _tot["sold"]

    # Header SUMAR (merged A:G)
    ws.merge_cells(f"A{row}:G{row}")
    c = ws[f"A{row}"]
    c.value = f"📊 SUMAR FINANCIAR — {title_period}"
    c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
    c.fill = header_fill
    c.alignment = center
    c.border = header_border
    ws.row_dimensions[row].height = 30
    row += 1

    # Rând 1: TOTAL ÎNCASĂRI (verde)
    ws.merge_cells(f"A{row}:D{row}")
    for col in ["A", "B", "C", "D"]:
        cell = ws[f"{col}{row}"]
        cell.fill = income_fill
        cell.border = thin_border
    cA = ws[f"A{row}"]
    cA.value = "💰 TOTAL ÎNCASĂRI"
    cA.font = Font(name="Calibri", bold=True, size=12, color="1F6B2A")
    cA.alignment = left_indent

    ws.merge_cells(f"E{row}:G{row}")
    for col in ["E", "F", "G"]:
        cell = ws[f"{col}{row}"]
        cell.fill = income_fill
        cell.border = thin_border
    cE = ws[f"E{row}"]
    cE.value = total_incasari_all
    cE.font = Font(name="Calibri", bold=True, size=12, color="1F6B2A")
    cE.alignment = right
    cE.number_format = num_fmt

    ws.row_dimensions[row].height = 28
    row += 1

    # Rând 2: TOTAL CHELTUIELI (roșu)
    ws.merge_cells(f"A{row}:D{row}")
    for col in ["A", "B", "C", "D"]:
        cell = ws[f"{col}{row}"]
        cell.fill = expense_fill
        cell.border = thin_border
    cA = ws[f"A{row}"]
    cA.value = "💸 TOTAL CHELTUIELI"
    cA.font = Font(name="Calibri", bold=True, size=12, color="C00000")
    cA.alignment = left_indent

    ws.merge_cells(f"E{row}:G{row}")
    for col in ["E", "F", "G"]:
        cell = ws[f"{col}{row}"]
        cell.fill = expense_fill
        cell.border = thin_border
    cE = ws[f"E{row}"]
    cE.value = total_plati_all
    cE.font = Font(name="Calibri", bold=True, size=12, color="C00000")
    cE.alignment = right
    cE.number_format = num_fmt

    ws.row_dimensions[row].height = 28
    row += 1

    # Rând 3: PROFIT NET / PIERDERE (font 14pt, fundal galben/roșu)
    if profit_net >= 0:
        profit_color = "1F6B2A"
        profit_fill_pat = profit_positive_fill
        profit_emoji = "💎"
        profit_label = "PROFIT NET (cash flow)"
    else:
        profit_color = "C00000"
        profit_fill_pat = profit_negative_fill
        profit_emoji = "⚠️"
        profit_label = "PIERDERE (cash flow)"

    ws.merge_cells(f"A{row}:D{row}")
    for col in ["A", "B", "C", "D"]:
        cell = ws[f"{col}{row}"]
        cell.fill = profit_fill_pat
        cell.border = header_border
    cA = ws[f"A{row}"]
    cA.value = f"{profit_emoji} {profit_label}"
    cA.font = Font(name="Calibri", bold=True, size=14, color=profit_color)
    cA.alignment = left_indent

    ws.merge_cells(f"E{row}:G{row}")
    for col in ["E", "F", "G"]:
        cell = ws[f"{col}{row}"]
        cell.fill = profit_fill_pat
        cell.border = header_border
    cE = ws[f"E{row}"]
    cE.value = profit_net
    cE.font = Font(name="Calibri", bold=True, size=14, color=profit_color)
    cE.alignment = right
    cE.number_format = num_fmt

    ws.row_dimensions[row].height = 36
    row += 2

    # ── Notă explicativă deductibilitate ────────────────────────────────
    ws.merge_cells(f"A{row}:G{row}")
    c = ws[f"A{row}"]
    c.value = (
        "ℹ️ Profitul deductibil fiscal (pentru ANAF) poate diferi: anumite "
        "cheltuieli auto/telecom sunt deductibile parțial (50%). "
        "Vezi raportul lunar al botului pentru detalii fiscale."
    )
    c.font = small
    c.alignment = center
    ws.row_dimensions[row].height = 22
    row += 2

    # ── FIX BUG #3: Footer alignment consistent (A:D + E:G) ─────────────
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"].value = (
        f"Data întocmirii: {datetime.now().strftime('%d.%m.%Y')}"
    )
    ws[f"A{row}"].font = small
    ws[f"A{row}"].alignment = Alignment(
        horizontal="left", vertical="center", indent=1
    )

    ws.merge_cells(f"E{row}:G{row}")
    ws[f"E{row}"].value = "Semnătura titularului: ___________________"
    ws[f"E{row}"].font = small
    ws[f"E{row}"].alignment = right
    ws.row_dimensions[row].height = 20
    row += 1

    # Disclaimer
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"].value = (
        "⚠️ Document generat automat de Bot Contabil PFA conform OMFP "
        "170/2015 (cod 14-1-1/b). Verificați cu contabilul autorizat "
        "înainte de depunere oficială."
    )
    ws[f"A{row}"].font = Font(
        name="Calibri", size=8, color="FF0000", italic=True
    )
    ws[f"A{row}"].alignment = center
    ws.row_dimensions[row].height = 18

    # ── Page setup ──────────────────────────────────────────────────────
    ws.freeze_panes = f"A{data_start_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.print_title_rows = f"1:{data_start_row - 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ============================================================
#              CSV FALLBACK (același 7-col format)
# ============================================================

def generate_registru_csv(
    transactions, year,
    pfa_name="PFA",
    pfa_cui="",
    month=None,
) -> bytes:
    """Fallback CSV — folosit dacă openpyxl nu e disponibil."""
    pfa_name, pfa_cui = _validate_pfa_info(pfa_name, pfa_cui)

    import csv
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";")

    period = (
        f"{LUNI_RO_UPPER.get(month, '')} {year}" if month
        else f"ANUL {year}"
    )
    writer.writerow([f"REGISTRU DE ÎNCASĂRI ȘI PLĂȚI — {period}"])
    writer.writerow([
        pfa_name,
        f"CUI: {pfa_cui}" if pfa_cui else "CUI: nesetat",
        "Sistem real de impunere",
        "cod 14-1-1/b",
    ])
    writer.writerow([])
    writer.writerow([
        "Nr.", "Data", "Document", "Explicații",
        "Încasări (RON)", "Plăți (RON)", "Sold cumulat (RON)"
    ])

    sold = 0.0
    nr = 0
    total_incasari = 0.0
    total_plati = 0.0
    relevant = sorted(
        [tx for tx in transactions
         if tx.tx_type in ("INCOME", "EXPENSE")
         and (month is None or
              (tx.occurred_on and tx.occurred_on.month == month))],
        key=lambda tx: (tx.occurred_on or date(year, 1, 1),
                        tx.id if tx.id else 0)
    )

    for tx in relevant:
        nr += 1
        tx_date = (
            tx.occurred_on.strftime("%d.%m.%Y") if tx.occurred_on else ""
        )
        is_income = tx.tx_type == "INCOME"
        amount = round(tx.amount_brut or 0.0, 2)

        if is_income:
            incasare = amount
            plata = 0.0
            sold = round(sold + incasare, 2)
            total_incasari += incasare
        else:
            incasare = 0.0
            plata = amount
            sold = round(sold - plata, 2)
            total_plati += plata

        writer.writerow([
            nr,
            tx_date,
            _resolve_document_type(tx),
            _build_explicatii(tx),
            f"{incasare:.2f}" if incasare else "",
            f"{plata:.2f}" if plata else "",
            f"{sold:.2f}",
        ])

    # SUMAR FINANCIAR la final
    total_incasari = round(total_incasari, 2)
    total_plati = round(total_plati, 2)
    profit_net = round(total_incasari - total_plati, 2)

    writer.writerow([])
    writer.writerow([f"=== SUMAR FINANCIAR — {period} ==="])
    writer.writerow([
        "TOTAL INCASARI", "", "", "", f"{total_incasari:.2f}", "", ""
    ])
    writer.writerow([
        "TOTAL CHELTUIELI", "", "", "", "", f"{total_plati:.2f}", ""
    ])
    profit_label = (
        "PROFIT NET (cash flow)" if profit_net >= 0
        else "PIERDERE (cash flow)"
    )
    writer.writerow([
        profit_label, "", "", "", "", "", f"{profit_net:.2f}"
    ])

    return buf.getvalue().encode("utf-8-sig")


def filename_registru(year, fmt="xlsx", month=None):
    if month:
        return f"registru_incasari_plati_{year}_{month:02d}.{fmt}"
    return f"registru_incasari_plati_{year}.{fmt}"
